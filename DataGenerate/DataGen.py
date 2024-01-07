import openpyxl


def dataGenerator():
    wk = openpyxl.load_workbook('TestData/TData10.xlsx')
    sh = wk['Sheet1']
    li = [] # outer list

    i = 1
    while sh.cell(i,1).value != None:
        li1=[] # inner list
        un = sh.cell(i,1)
        pw = sh.cell(i,2)
        li1.insert(0,un.value)
        li1.insert(1,pw.value)
        li.insert(i-1,li1)
        i += 1
    return li